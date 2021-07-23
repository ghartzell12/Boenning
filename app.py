from flask import Flask, render_template, request, redirect, url_for
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename
from openpyxl import Workbook
from threading import Thread
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import pandas as pd
# api via https://pypi.org/project/googlesearch-python/
from googlesearch import search
import requests
import bs4
from urllib.request import urlopen
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import phonenumbers
from datetime import date, datetime
import timeit
from multiprocessing import Process
import time
from func_timeout import func_timeout, FunctionTimedOut
from tkinter.filedialog import asksaveasfile
import os
import io
import smtplib
from gav import excel_formatter
import xlrd

app = Flask(__name__)


def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in xrange(0, nrows):
        for col in xrange(0, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    return book1


def run(query, req, keywords, results):
    query2 = query + "bank"
    for i in req.split(','):
        query2+='"'+ i +'"'
    print(query2)
    start_time = datetime.now()
    links = search(query2, num_results = results)
    wb = Workbook()
    ws = wb.active
    format_sheet(ws)
    for val in range(0,min(results, len(links))):
        print(links[val])
        keyword_count = 0
        keyword_desc = ""
        phone = []
        email = []
        notes = ""
        start = time.time()
        try:
            func_timeout(10, parse, args = (links[val],"",[],ws,val+2, keywords, keyword_count, keyword_desc, phone, email, notes))
        except:
            pass
    today = date.today()
    files = [('Excel Files', '*.xlsx'),
             ('All Files', '*.*')]
    end_time = datetime.now()
    runtime = end_time - start_time
    ws["H1"] = "Runtime: " + str(runtime.total_seconds())
    if 1 == 0:
        with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login('gavinhartzellboenning@gmail.com', "Redsox09")
            subject = 'Lender Tracker Research Run'
            body = 'See Attached.'
            msg = MIMEMultipart('Boenning')
            msg['Subject'] =  subject
            msg['From'] = 'gavinhartzellboenning@gmail.com'
            msg['To'] = 'gavin.hartzell42@gmail.com'
            part = MIMEBase('application', "octet-stream")
            df = pd.DataFrame(ws.values)
            towrite = io.BytesIO()
            df.to_excel(towrite)  # write to BytesIO buffer
            part.set_payload(towrite.getvalue())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment; filename= "Result.xlsx"')
            msg.attach(part)
            smtp.sendmail('gavinhartzellboenning@gmail.com','gavin.hartzell42@gmail.com', msg)
    wb.save(query+".xlsx")
def stopper():
    start = time.time()
    print(time.time)
    while time.time() < start + 15:
        pass
    print("exception")
    raise Exception("no time")
def format_sheet(ws):
    for letter in ['A','B','C','D','E','F','G']:
        redFill = PatternFill(start_color='ADD8E6',
                   end_color='ADD8E6',
                   fill_type='solid')
        ws[letter+'1'].fill = redFill
    gav = {
        "A1":"Title",
        "B1": "Link",
        "C1": "RC Notes",
        "D1" : "Email",
        "E1" : "Phone Numbers",
        "F1" : "Keywords Hit",
        "G1" :"Keyword Count"

    }
    for key in gav:
        ws[key] = gav[key]
        ws[key].font = Font(bold=True)

def parse(base, site, urls,ws,row, keywords, keyword_count, keyword_desc, phone, email, notes):
    keyword_count, keyword_desc, phone, email, notes = scrape(base+site,row, True,ws, keywords, keyword_count, keyword_desc, phone, email, notes)
    temporary_parse = requests.get(base+site)
    print(temporary_parse.status_code)
    ws['B'+str(row)] = base
    soup = bs4.BeautifulSoup(temporary_parse.text,"html.parser")
    for link in soup.find_all('a'):
        # get the href, if not searched then search it
        if 'href' in link:
            print(link)
            if link['href'] not in urls and 'html' in link['href']:
                urls.append(link['href'])
                print("new url")
                keyword_count, keyword_desc, phone, email, notes = scrape(base+link['href'], False,ws, keywords, keyword_count, keyword_desc, phone, email, notes)
                parse(base, link['href'], urls,ws,row,keywords, keyword_count, keyword_desc, phone, email, notes)

    if keyword_desc != "":
        ws["F"+str(row)] = keyword_desc
    else:
        ws["F"+str(row)] = "Key Words were not found"
    ws["G"+str(row)] = keyword_count
    email_string = ""
    for e in email:
        email_string += (e)
        email_string += " : "
    phone_string = ""
    for p in phone:
        phone_string += (p)
        phone_string += " : "
    if email_string != "":
        ws["D"+str(row)] = email_string
    else:
        ws["D"+str(row)] = "No Email Found"
    if phone_string != "":
        ws["E"+str(row)] = phone_string
    else:
        ws["E"+str(row)] = "None Found"
    if notes != "":
        ws["C" + str(row)] = notes
    else:
        ws["C" + str(row)] = "None Found"

import html2text
from openpyxl.styles import PatternFill


def scrape(site,row, title_change,ws, keywords, keyword_count, keyword_desc, phone, email, notes):
    # extract title
    temporary_parse = requests.get(site)
    soup = bs4.BeautifulSoup(temporary_parse.text,"html.parser")
    text = soup.get_text()
    title = soup.find('title')
    try:
        print(title.get_text())
        print('here')
        if title_change:
            ws['A'+str(row)] = title.get_text()
    except:
        print("Not Found")
    for keyword in keywords:
        if keyword in text:
            keyword_desc += " | " + text[text.index(keyword):text.index(keyword)+len(keyword)]
            keyword_count += 1
    emails = []
    emails += re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", text)
    mailtos = soup.select('a[href^=mailto]')
    for i in mailtos:
        href=i['href']
        try:
            str1, str2 = href.split(':')
        except ValueError:
            break

        emails.append(str2)
    phone = phone + (re.findall("[(][\d]{3}[)][ ]?[\d]{3}-[\d]{4}", text))
    print(phone)
    if "LTV" in text:
        notes += (text[text.index("LTV")-40:text.index("LTV") +40]) + " | "
    if "term" in text:
        notes += (text[text.index("term")-40:text.index("term") +40]) + " | "
    if "year" in text:
        notes += (text[text.index("year")-40:text.index("year") +40]) + " | "
    if "LTC" in text:
        notes +=  text[text.index("LTC")-40:text.index("LTC") + 40] + " | "
    if "$" in text:
        notes +=  text[text.index("$") - 40:text.index("$") + 40] + " | "
    if "HUD" in text:
        notes +=  text[text.index("HUD") - 40:text.index("HUD") + 40] + " | "
    if "FHA" in text:
        notes +=  text[text.index("FHA") - 40:text.index("FHA") + 40] + " | "
    if "debt service" in text:
        notes +=  text[text.index("debt service") - 40:text.index("debt service") + 40] + " | "
    if "loan to" in text:
        notes +=  text[text.index("loan to") - 40:text.index("loan to") + 40] + " | "
    email = email + emails
    return keyword_count, keyword_desc, phone, email, notes
    # other possible contains
    # extract LTV, LTC, anyting else relevant

@app.route("/")
def home():
    return(render_template("index.html"))

@app.route("/complete/")
def complete():
    return(render_template("complete.html", name = "Blub"))



@app.route("/lender/", methods = ["POST","GET"])
def lender():
    if request.method == "POST":
        print("posted")
        query = request.form["query"]
        keywords = request.form["key"]
        req = request.form["req"]
        q = int(request.form["quant"])
        run(query, req, keywords.split(","),q)
        return (render_template("complete.html", name = query+".xlsx"))
    return(render_template("rcformat.html"))

@app.route("/rc/", methods = ["POST","GET"])
def rc():
    if request.method == "POST":
        print("posted")
        ev_complete_raw = request.form["evc"]
        ltv1_raw = request.form["ltv1"]
        ev_stable_raw = request.form["evs"]
        ltv2_raw = request.form["ltv2"]
        ltc_raw = request.form["ltc"]
        noi_raw = request.form["noi"]
        year_raw = request.form["year"]
        dscr_raw = request.form["dscr"]
        try:
            ev_complete = float(ev_complete_raw)
            if ev_complete == None:
                ev_complete = 0
        except:
            ev_complete = 0

        try:
            ltv1 = str(ltv1_raw)
            if ltv1_raw == None:
                ltv1 = "nnnn"
        except:
            ltv1 = "nnn"

        try:
            ltc = str(ltc_raw)
            if ltc_raw == None:
                ltc = "nnnn"
        except:
            ltc = "nnn"

        try:
            ev_stable = float(ev_stable_raw)
            if ev_stable == None:
                ev_stable = 0
        except:
            ev_stable = 0

        try:
            ltv2 = str(ltv2_raw)
            if ltv2_raw == None:
                ltv2 = "nnnn"
        except:
            ltv2 = "nnn"

        try:
            noi = float(noi_raw)
            if noi == None:
                noi = 0
        except:
            noi = 0

        try:
            year = int(year_raw)
            if year == None:
                year = 0
        except:
            year = 0

        try:
            dscr = str(dscr_raw)
            if dscr_raw == None:
                dscr = "nnn"
        except:
            dscr = "nnn"
        print("yes")
        if request.files:
            print("IF")
            filename = 'boenning'
            uploaded_file = request.files["filename"].read()
            wb = load_workbook(open_xls_as_xlsx(uploaded_file))
        else:
            return(render_template("otherproj.html"))
        print("here")
        excel_formatter(wb,ev_complete,ltv1,ev_stable,ltv2,ltc,noi,year,dscr, filename)
        return (render_template("complete.html", name = "boenning.xlsx"))
    return(render_template("otherproj.html"))

if __name__ == "__main__":
    app.run()

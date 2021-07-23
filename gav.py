from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename, askopenfile
from tkinter.filedialog import asksaveasfilename, asksaveasfile
from openpyxl import Workbook, load_workbook
from threading import Thread
from openpyxl.styles import Border, Side, Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
# api via https://pypi.org/project/googlesearch-python/
from googlesearch import search
import requests
import bs4
from urllib.request import urlopen
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import phonenumbers
from datetime import date, datetime
import timeit
from multiprocessing import Process
import time
from func_timeout import func_timeout, FunctionTimedOut
import openpyxl
import os

def excel_formatter(wb, ev_complete, ltv1, ev_stable, ltv2, ltc, noi, year, dscr, filename):
    sources_exists = False
    debt_service_exists = False
    try:
        sources = wb[wb.sheetnames[0]]
        sources_exists = True
    except:
        sources = wb.create_sheet('Debt Service')
    debt_service = wb[wb.sheetnames[1]]
    debt_service_exists = True
    #except:
        #debt_service = wb.create_sheet('Debt Service')
    celldict = {}
    if sources_exists:
        for row in range(1,len(sources['B'])):
            sources['B' + str(row)].border = Border(left = Side(border_style='thin', color='FF000000'))
            sources['F' + str(row)].border = Border(right = Side(border_style='thin', color='FF000000'))
        bro = 0
        for i in range(1,55):
            if '|' in str(sources['C'+str(i)].value):
                sources['C'+str(i)] = ''
            if 'dated' in str(sources['C'+str(i)].value):
                sources['C'+str(i)] = ''
            if "good" in str(sources['C'+str(i)].value):
                bro = i
        if bro != 0:
            for c in 'BCDEF':
                for row in [bro,bro+1,bro-1, bro+2]:
                    sources[c+str(row)].fill = PatternFill(bgColor="0d0080", fgColor="0d0080",fill_type = "solid")
        for col in 'CDE':
            sources[col + '5'].border = Border(bottom = Side(border_style='thin', color='FF000000'))
            sources[col + '11'].border = Border(bottom = Side(border_style='thin', color='FF000000'))
        for col in 'CDE':
            for row in range(14,bro-3):
                if sources['C'+str(row)].value != None and str(sources['C'+str(row)].value) != " ":
                    if "Sources" not in sources['C'+str(row)].value and "Use" not in sources['C'+str(row)].value:
                        sources[col + str(row)].border = Border(bottom = Side(border_style='thin', color='00808080'))
                    else:
                        sources[col + str(row)].border = Border(bottom = Side(border_style='thin', color='00000000'))
        sources.column_dimensions['E'].width = 24
        for i in range(6,10):
            sources.row_dimensions[i].height = 18
        sources.row_dimensions[33].height = 20
        for c in 'BCDEF':
            for row in [1,2]:
                sources[c+str(row)].fill = PatternFill(bgColor="0d0080", fgColor="0d0080",fill_type = "solid")
        contingency = 0
        costsrow = 17
        totalsourcesrow = 0
        titlefont = Font(name='Times New Roman',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='0d0080')
        valuefont = Font(name='Times New Roman',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        for i in range(10,30):
            if "Deposit to" in str(sources["C"+str(i)].value):
                celldict["Project Cost"] = "C" + str(i)
            if "Total Source" in str(sources["C"+str(i)].value):
                totalsourcesrow = i
            if "Par Amount" in str(sources["C"+str(i)].value):
                celldict["Senior Debt"] =  "C" + str(i)
            if "Contingency" in str(sources["C"+str(i)].value):
                celldict["Contingency"] =  "C" + str(i)
            if "Costs of" in str(sources["C"+str(i)].value):
                celldict["Costs"] =  "C" + str(i)
                costsrow = i
            if "Round" in str(sources["C"+str(i)].value):
                contingency = float(sources["E"+str(i)].value)
                sources.delete_rows(i)
        sources["E" + str(costsrow)] = float(sources["E" + str(costsrow)].value) + contingency
        if "Senior Debt" in celldict:
            sources[celldict["Senior Debt"]] = "Senior Debt"
        if "Project Cost" in celldict:
            sources[celldict["Project Cost"]] = "Project Cost"
        if "Costs" in celldict:
            sources[celldict['Costs']] = "Financing Fees & Closing Costs"
        count = 0
        for a in [ltv1, ltv2]:
            if "y" in a:
                count +=2
        if "y" in ltc:
            count+=1
        running_count = 6
        if count == 5:
            running_count = 6
        else:
            running_count = 7
        for col in "EF":
            for row in range(6,11):
                sources[col+str(row)] = ''
        if "y" in ltv1 and ev_complete > 0:
            sources["E"+str(running_count)] = "Estimated Value (Complete)"
            sources["E"+str(running_count)].font = titlefont
            sources["F" + str(running_count)] = ev_complete
            sources["E"+str(running_count)].font = titlefont
            sources["F"+str(running_count)].font = valuefont
            sources["E"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000') )
            sources["F"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000'))
            sources["F"+str(running_count)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
            running_count += 1
            sources["E"+str(running_count)] = "LTV (Complete)"
            sources["F" + str(running_count)] = '=C5/F'+str(running_count-1)
            sources["E"+str(running_count)].font = titlefont
            sources["F"+str(running_count)].font = valuefont
            sources["E"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000') )
            sources["F"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000'))
            sources["F"+str(running_count)].number_format = '0.00%'
            running_count +=1
            #Find the best cell to put this in, put it there
        if "y" in ltv2 and ev_stable > 0:
            sources["E"+str(running_count)] = "Estimated Value (Stabilized)"
            sources["E"+str(running_count)].font = titlefont
            sources["F" + str(running_count)] = ev_complete
            sources["E"+str(running_count)].font = titlefont
            sources["F"+str(running_count)].font = valuefont
            sources["E"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000') )
            sources["F"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000'))
            sources["F"+str(running_count)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
            running_count += 1
            sources["E"+str(running_count)] = "LTV (Stabilized)"
            sources["E"+str(running_count)].font = valuefont
            sources["F" + str(running_count)] = '=C5/F'+str(running_count-1)
            sources["E"+str(running_count)].font = titlefont
            sources["F"+str(running_count)].font = valuefont
            sources["E"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000') )
            sources["F"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000'))
            sources["F"+str(running_count)].number_format = '0.00%'
            running_count +=1
            #Find the best cell to put tis in, put it there
        if "y" in ltc:
            sources["E"+str(running_count)] = "LTC"
            sources["E"+str(running_count)].font = titlefont
            sources["F" + str(running_count)] = '=C5/E' + str(totalsourcesrow)
            sources["E"+str(running_count)].font = valuefont
            sources["E"+str(running_count)].font = titlefont
            sources["F"+str(running_count)].font = valuefont
            sources["E"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000') )
            sources["F"+str(running_count)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                           right = Side(border_style='thin', color='FF000000'),
                                                           top = Side(border_style='thin', color='FF000000'),
                                                           bottom =Side(border_style='thin', color='FF000000'))
            sources["F"+str(running_count)].number_format = '0.00%'
            running_count+=1




    if debt_service_exists:
        debt_service['B5'].border = Border(bottom = Side(border_style=None, color='FF000000'))
        titlefont = Font(name='Times New Roman',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='0d0080')
        for row in range(1,len(debt_service['B'])):
            debt_service['B' + str(row)].border = Border(left = Side(border_style='thin', color='FF000000'))
            debt_service['N' + str(row)].border = Border(right = Side(border_style='thin', color='FF000000'))
        debt_service['B11'].border = Border(bottom = Side(border_style='thin', color='FF000000'))
        for col in 'CDE':
            debt_service[col + '5'].border = Border(bottom = Side(border_style='thin', color='FF000000'))
            debt_service[col + '11'].border = Border(bottom = Side(border_style='thin', color='FF000000'))
        for col in 'CDE':
            for row in range(14,30):
                if str(debt_service['C'+str(row)].value) == '0':
                    debt_service[col + str(row)].border = Border(bottom = Side(border_style='thin', color='00808080'),
                                                                top = Side(border_style='thin', color='00808080'))
                if debt_service['C'+str(row)].value != None and str(debt_service['C'+str(row)].value) != " " and "Scatter" not in str(debt_service['C'+str(row)].value) and "Public" not in str(debt_service['C'+str(row)].value):
                    debt_service[col + str(row)].border = Border(bottom = Side(border_style='thin', color='00808080'))
        debt_service["C10"] = "Debt Service Schedule"
        debt_service["G14"] = "Interest Rate"
        debt_service.column_dimensions['G'].width = 12.5
        debt_service.column_dimensions['L'].width = 12.5
        debt_service.column_dimensions['M'].width = 6
        debt_service.delete_cols(13,1)
        bluerow = 1
        debt_service.insert_cols(12)
        year_row = 0
        for col in 'CDEFGHIJKLM':
            debt_service[col + '5'].border = Border(bottom = Side(border_style='thin', color='FF000000'))
            debt_service[col + '11'].border = Border(bottom = Side(border_style='thin', color='FF000000'))
        # clear col M or borders
        for i in range(14,80):
            debt_service["M"+str(i)].border = Border(bottom = Side(border_style=None, color='FF000000'),
                                                    top = Side(border_style=None, color='FF000000'))
            borders = {}
        for c in 'BCDEFGHIJKLMN':
            for row in [1,2]:
                debt_service[c+str(row)].fill = PatternFill(bgColor="0d0080", fgColor="0d0080",fill_type = "solid")
        for i in range(1,90):
            if '|' in str(debt_service['C'+str(i)].value):
                debt_service['C'+str(i)] = ''
            if 'ated' in str(debt_service['C'+str(i)].value):
                debt_service['C'+str(i)] = ''
            if "good" in str(debt_service['C'+str(i)].value):
                bluerow = i-1
                break
        for row_add in range(0,5):
            for col in "ABCDEFGHIJKLMN":
                debt_service[col + str(row_add+bluerow)].fill = PatternFill(bgColor="0d0080", fgColor="0d0080",fill_type = "solid")
        row = 1
        for row in range(0,100):
            if debt_service["C"+str(17+row)].value == None:
                break
            if (row+1) % 5 == 0:
                for col in "CDEFGHIJK":
                    debt_service[col+str(17+row)].border = Border(bottom = Side(border_style='thin', color='FF000000'))
                row += 1
            else:
                for col in "CDEFGHIJK":
                    debt_service[col+str(17+row)].border = Border(bottom = Side(border_style=None, color='FF000000'))
                row += 1
        if "y" in dscr and year > 0 and noi > 0:
            year_row = 0
            for row in range(17,30):
                year_str = "Nope"
                try:
                    year_str = str(debt_service["C"+str(row)].value.year)
                except:
                    year_str = "Nope"
                if year_str == str(year):
                    year_row = row
            if year_row != 0:
                debt_service["L14"] = "NOI"
                debt_service["M14"] = "DSCR"
                debt_service["L" + str(year_row)] = noi
                debt_service["M" + str(year_row)] = "=L"+str(year_row)+ "/I" +str(year_row)
                debt_service["M" + str(year_row)].number_format = '0.00'
                print(year_row)
                for c in "KLM":
                    debt_service[c + "14"].border = Border(top = Side(border_style='thin', color='FF000000'),
                                                          bottom = Side(border_style='thin', color='FF000000'))
                for r in range(15,year_row):
                    debt_service["K"+str(r)].border = Border(left = Side(border_style='thin', color='FF000000'))
                    debt_service["M"+str(r)].border = Border(right = Side(border_style='thin', color='FF000000'))
                debt_service["K14"].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                   top = Side(border_style='thin', color='FF000000'),
                                                   bottom = Side(border_style='thin', color='FF000000'))
                debt_service["M14"].border = Border(right = Side(border_style='thin', color='FF000000'),
                                                   top = Side(border_style='thin', color='FF000000'),
                                                   bottom = Side(border_style='thin', color='FF000000'))
                debt_service["K"+str(year_row)].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                   bottom = Side(border_style='thin', color='FF000000'))
                debt_service["L"+str(year_row)].border = Border(bottom = Side(border_style='thin', color='FF000000'))
                debt_service["M"+str(year_row)].border = Border(right = Side(border_style='thin', color='FF000000'),
                                                   bottom = Side(border_style='thin', color='FF000000'))
                debt_service["K14"].border = Border(left = Side(border_style='thin', color='FF000000'),
                                                   top = Side(border_style='thin', color='FF000000'),
                                                   bottom = Side(border_style='thin', color='FF000000'))
                titlefont = Font(name='Arial',size=9,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='0d0080')
                valuefont = Font(name='Times New Roman',size=8,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
                debt_service["L"+str(year_row)].font = valuefont
                debt_service["L14"].font = titlefont
                debt_service["L"+str(year_row)].alignment = Alignment(horizontal='right')
                debt_service["L14"].alignment = Alignment(horizontal='right')
        totalrow = 0
        for row in range(18,100):
            if str(debt_service["C"+str(row)].value) == '0':
                totalrow = row
                print(totalrow)
        if totalrow > 0:
            for c in "CDEFGHIJKLM":
                debt_service[c + str(totalrow)].border = Border(top = Side(border_style='thin', color='FF000000'),
                                       bottom = Side(border_style='thin', color='FF000000'))


    wb.save(os.path.expanduser("~")+"/Downloads/"+filename+'.xlsx')
